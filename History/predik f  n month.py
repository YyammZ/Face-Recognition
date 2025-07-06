import cv2
import torch
import joblib
import numpy as np
import threading
import time
import os
from datetime import datetime
from facenet_pytorch import MTCNN, InceptionResnetV1
from torchvision import transforms
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import mediapipe as mp
from collections import deque

# === Load Model & Encoder ===
model_data = joblib.load('model.pkl')
model = model_data['model']
label_encoder = model_data['label_encoder']

# === MTCNN & ResNet ===
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
mtcnn = MTCNN(keep_all=True, device=device)
resnet = InceptionResnetV1(pretrained='vggface2').eval().to(device)
normalize = transforms.Normalize([0.5], [0.5])

# === Shared State ===
latest_frame = [None]
frame_lock = threading.Lock()
last_detection = []
already_marked = set()
dicatat_hari_ini = set()

# === Config ===
PROCESS_EVERY_N_FRAMES = 2
UNKNOWN_THRESHOLD = 0.55
EXCEL_PATH = 'attendance.xlsx'
JAM_TEPAT_WAKTU = datetime.strptime("07:30", "%H:%M").time()

# === Excel Absensi ===
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
fill_hijau = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fill_merah = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def tulis_kehadiran_excel(nama, jam_datang):
    now = datetime.now()
    bulan_tahun = now.strftime("%B %Y")  # Contoh: "June 2025"
    hari_tanggal = now.strftime("%a/%d")
    tanggal_str = now.strftime("%d")
    key_hari_ini = f"{nama}-{hari_tanggal}"

    if key_hari_ini in dicatat_hari_ini:
        return

    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = bulan_tahun
        wb.save(EXCEL_PATH)

    wb = load_workbook(EXCEL_PATH)
    if bulan_tahun not in wb.sheetnames:
        ws = wb.create_sheet(title=bulan_tahun)
    else:
        ws = wb[bulan_tahun]
    # Cek apakah header sudah dibuat
    if not ws.cell(row=1, column=3).value:  # Jika belum ada tulisan "Face Recognition"
        # Header baris 1
        ws.merge_cells('C1:AG1')
        header_cell = ws.cell(row=1, column=3)
        header_cell.value = "Face Recognition"
        header_cell.font = Font(size=26, bold=True)
        header_cell.alignment = Alignment(horizontal='center')
    
        # Bulan
        ws['A2'] = bulan_tahun
        ws['A2'].font = Font(size=14, bold=True)
    
        # Header kolom NO dan Nama
        ws.merge_cells('A3:A4')
        cell = ws['A3']
        cell.value = "NO"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
    
        ws.merge_cells('B3:B4')
        cell = ws['B3']
        cell.value = "Nama"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
    
        # Header tanggal
        ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=33)
        cell = ws.cell(row=3, column=3)
        cell.value = "Waktu Kedatangan"
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
    
        # Header keterangan
        ws.merge_cells(start_row=3, start_column=34, end_row=3, end_column=35)
        cell = ws.cell(row=3, column=34)
        cell.value = "Keterangan"
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
    
        # Baris tanggal (Sen/01, Sel/02, dst)
        for i in range(31):
            try:
                tanggal = datetime(now.year, now.month, i + 1)
                kolom = i + 3
                cell = ws.cell(row=4, column=kolom)
                cell.value = tanggal.strftime("%a/%d")
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True)
            except:
                break
            
        # Keterangan On Time / Late
        ws.cell(row=4, column=34).value = "On Time"
        ws.cell(row=4, column=35).value = "Late"
        for col in [34, 35]:
            cell = ws.cell(row=4, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True)
    
        # Tambahkan border ke seluruh header
        for row in range(3, 5):
            for col in range(1, 36):
                ws.cell(row=row, column=col).border = thin_border

    # Cek baris untuk nama
    row = 5
    while True:
        nama_cell = ws.cell(row=row, column=2).value
        if nama_cell is None:
            ws.cell(row=row, column=1).value = row - 4
            ws.cell(row=row, column=2).value = nama
            break
        elif nama_cell == nama:
            break
        row += 1

    # Cari kolom untuk tanggal hari ini
    for col in range(3, 34):
        header = ws.cell(row=4, column=col).value
        if header and tanggal_str in header:
            if ws.cell(row=row, column=col).value:
                return  # Sudah dicatat

            jam_kehadiran = datetime.now().strftime("%H:%M")
            cell = ws.cell(row=row, column=col)
            cell.value = jam_kehadiran
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            jam_obj = datetime.strptime(jam_kehadiran, "%H:%M").time()

            if jam_obj <= JAM_TEPAT_WAKTU:
                nilai = ws.cell(row=row, column=34).value or 0
                ws.cell(row=row, column=34).value = nilai + 1
                cell.fill = fill_hijau
            else:
                nilai = ws.cell(row=row, column=35).value or 0
                ws.cell(row=row, column=35).value = nilai + 1
                cell.fill = fill_merah

            ws.cell(row=row, column=34).border = thin_border
            ws.cell(row=row, column=35).border = thin_border
            break

    # Tambahkan border ke seluruh baris baru
    for col in range(1, 36):
        ws.cell(row=row, column=col).border = thin_border

    wb.save(EXCEL_PATH)
    dicatat_hari_ini.add(key_hari_ini)

# === MediaPipe Face Detection & FaceMesh ===
mp_face_detection = mp.solutions.face_detection
mp_drawing = mp.solutions.drawing_utils
face_detector = mp_face_detection.FaceDetection(model_selection=1, min_detection_confidence=0.6)

mp_face_mesh = mp.solutions.face_mesh
face_mesh = mp_face_mesh.FaceMesh(static_image_mode=False, max_num_faces=1, refine_landmarks=True)

blink_buffer = deque(maxlen=5)
blink_timestamps = {}

def calculate_ear(eye_landmarks):
    A = np.linalg.norm(np.array(eye_landmarks[1]) - np.array(eye_landmarks[5]))
    B = np.linalg.norm(np.array(eye_landmarks[2]) - np.array(eye_landmarks[4]))
    C = np.linalg.norm(np.array(eye_landmarks[0]) - np.array(eye_landmarks[3]))
    ear = (A + B) / (2.0 * C)
    return ear

def detect_blink(frame, name):
    h, w, _ = frame.shape
    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    results = face_mesh.process(rgb)
    if not results.multi_face_landmarks:
        return False
    for face_landmarks in results.multi_face_landmarks:
        mesh_points = [(int(p.x * w), int(p.y * h)) for p in face_landmarks.landmark]
        left_eye_idx = [362, 385, 387, 263, 373, 380]
        right_eye_idx = [33, 160, 158, 133, 153, 144]
        left_eye = [mesh_points[i] for i in left_eye_idx]
        right_eye = [mesh_points[i] for i in right_eye_idx]
        left_ear = calculate_ear(left_eye)
        right_ear = calculate_ear(right_eye)
        ear = (left_ear + right_ear) / 2.0
        if ear < 0.2:
            blink_timestamps[name] = time.time()
            return True
    return False

def recent_blink(name, threshold=2.0):
    return name in blink_timestamps and (time.time() - blink_timestamps[name]) < threshold

# === Camera Thread ===
def camera_reader():
    cap = cv2.VideoCapture(0)
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    while True:
        ret, frame = cap.read()
        if not ret:
            continue
        frame = cv2.flip(frame, 1)
        with frame_lock:
            latest_frame[0] = frame.copy()
        time.sleep(0.005)

# === Face Recognition Thread ===
def face_recognition_worker():
    global last_detection
    frame_count = 0
    while True:
        with frame_lock:
            frame = latest_frame[0]
        if frame is None:
            time.sleep(0.01)
            continue

        if frame_count % PROCESS_EVERY_N_FRAMES == 0:
            try:
                rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                results = face_detector.process(rgb)
                result = []

                if results.detections:
                    for detection in results.detections:
                        bboxC = detection.location_data.relative_bounding_box
                        ih, iw, _ = frame.shape
                        x1 = int(bboxC.xmin * iw)
                        y1 = int(bboxC.ymin * ih)
                        w_box = int(bboxC.width * iw)
                        h_box = int(bboxC.height * ih)

                        # Perbesar bounding box
                        scale = 1.5
                        cx = x1 + w_box // 2
                        cy = y1 + h_box // 2
                        w_scaled = int(w_box * scale)
                        h_scaled = int(h_box * scale)

                        x1 = max(0, cx - w_scaled // 2)
                        y1 = max(0, cy - h_scaled // 2)
                        x2 = min(iw, cx + w_scaled // 2)
                        y2 = min(ih, cy + h_scaled // 2)

                        face = frame[y1:y2, x1:x2]
                        if face.size == 0 or face.shape[0] < 60 or face.shape[1] < 60:
                            continue

                        face = cv2.resize(face, (160, 160), interpolation=cv2.INTER_CUBIC)
                        tensor = torch.from_numpy(face / 255.).permute(2, 0, 1).float()
                        tensor = normalize(tensor).unsqueeze(0).to(device)

                        with torch.no_grad():
                            embedding = resnet(tensor).cpu().numpy()

                        pred = model.predict(embedding)[0]
                        prob = model.predict_proba(embedding)[0]
                        confidence = np.max(prob)
                        name = label_encoder.inverse_transform([pred])[0] if confidence >= UNKNOWN_THRESHOLD else "Gk Kenal"

                        result.append((x1, y1, x2, y2, name, confidence))

                        if name != "Gk Kenal":
                            if detect_blink(frame, name) and recent_blink(name):
                                jam_datang = datetime.now().strftime("%H:%M")
                                tulis_kehadiran_excel(name, jam_datang)

                last_detection = result
            except Exception as e:
                print("Recognition error:", e)

        frame_count += 1
        time.sleep(0.01)

# === Display Thread ===
def display_loop():
    window_name = 'Face Recognition (Absensi)'
    cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
    is_fullscreen = False  # Status fullscreen

    while True:
        with frame_lock:
            frame = latest_frame[0].copy() if latest_frame[0] is not None else None
        if frame is None:
            continue

        for x1, y1, x2, y2, name, conf in last_detection:
            box_color = (0, 255, 0) if name != "Gk Kenal" else (0, 0, 255)  # Tetap hijau untuk dikenal, merah untuk tidak dikenal
            text_color = (0, 255, 255)  # Kuning
            outline_color = (0, 0, 0)   # Hitam (outline teks)

            # Kotak wajah
            cv2.rectangle(frame, (x1, y1), (x2, y2), box_color, 2)

            # Label teks
            label = f"{name} ({conf:.2f})"

            # Outline teks (biar terbaca lebih baik)
            cv2.putText(frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, outline_color, 3)
            # Teks utama (kuning)
            cv2.putText(frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, text_color, 1)

        cv2.imshow(window_name, frame)

        key = cv2.waitKey(1) & 0xFF
        if key == ord('q'):
            break
        elif key == ord('f'):
            is_fullscreen = not is_fullscreen
            if is_fullscreen:
                cv2.setWindowProperty(window_name, cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)
            else:
                cv2.setWindowProperty(window_name, cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_NORMAL)

    cv2.destroyAllWindows()


# === Run Threads ===
threading.Thread(target=camera_reader, daemon=True).start()
threading.Thread(target=face_recognition_worker, daemon=True).start()
display_loop()
