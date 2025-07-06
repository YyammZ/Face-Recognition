import cv2
import torch
import joblib
import numpy as np
import threading
import time
import os
from datetime import datetime
from facenet_pytorch import MTCNN, InceptionResnetV1
from torchvision import transforms
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import mediapipe as mp
from collections import deque
import requests
import json
import pygame
import threading

# === Load Model & Encoder ===
model_data = joblib.load('model.pkl')
model = model_data['model']
label_encoder = model_data['label_encoder']

# === MTCNN & ResNet ===
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
mtcnn = MTCNN(keep_all=True, device=device)
resnet = InceptionResnetV1(pretrained='vggface2').eval().to(device)
normalize = transforms.Normalize([0.5], [0.5])

# === Shared State ===
latest_frame = [None]
frame_lock = threading.Lock()
last_detection = []
already_marked = set()
dicatat_hari_ini = set()
foto_dicatat_hari_ini = set()
telegram_dicatat_hari_ini = set() 
running = True
recognized_since = {}
notifikasi_text = ""
notifikasi_timer = 0

# === Config ===
PROCESS_EVERY_N_FRAMES = 2
UNKNOWN_THRESHOLD = 0.55
EXCEL_PATH = 'attendance.xlsx'
JAM_TEPAT_WAKTU = datetime.strptime("07:30", "%H:%M").time()

# === Excel Absensi ===
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
fill_hijau = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fill_merah = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def tulis_kehadiran_excel(nama, jam_datang, chat_id=None):
    now = datetime.now()
    bulan_tahun = now.strftime("%B %Y")
    hari_tanggal = now.strftime("%a/%d")
    tanggal_str = now.strftime("%d")
    key_hari_ini = f"{nama}-{hari_tanggal}"

    if key_hari_ini in dicatat_hari_ini:
        return False  # Tidak dicatat ulang

    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = bulan_tahun
        wb.save(EXCEL_PATH)

    wb = load_workbook(EXCEL_PATH)
    ws = wb[bulan_tahun] if bulan_tahun in wb.sheetnames else wb.create_sheet(title=bulan_tahun)

    # Cek apakah header sudah dibuat
    if not ws.cell(row=1, column=3).value:
        ws.merge_cells('C1:AG1')
        header_cell = ws.cell(row=1, column=3)
        header_cell.value = "Face Recognition"
        header_cell.font = Font(size=26, bold=True)
        header_cell.alignment = Alignment(horizontal='center')

        ws['A2'] = bulan_tahun
        ws['A2'].font = Font(size=14, bold=True)

        ws.merge_cells('A3:A4')
        cell = ws['A3']
        cell.value = "NO"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)

        ws.merge_cells('B3:B4')
        cell = ws['B3']
        cell.value = "Nama"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)

        ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=33)
        cell = ws.cell(row=3, column=3)
        cell.value = "Waktu Kedatangan"
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)

        ws.merge_cells(start_row=3, start_column=34, end_row=3, end_column=35)
        cell = ws.cell(row=3, column=34)
        cell.value = "Keterangan"
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)

        for i in range(31):
            try:
                tanggal = datetime(now.year, now.month, i + 1)
                kolom = i + 3
                cell = ws.cell(row=4, column=kolom)
                cell.value = tanggal.strftime("%a/%d")
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True)
            except:
                break

        ws.cell(row=4, column=34).value = "On Time"
        ws.cell(row=4, column=35).value = "Late"
        for col in [34, 35]:
            cell = ws.cell(row=4, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True)

        for row in range(3, 5):
            for col in range(1, 36):
                ws.cell(row=row, column=col).border = thin_border

    # Cek baris nama
    row = 5
    while True:
        nama_cell = ws.cell(row=row, column=2).value
        if nama_cell is None:
            ws.cell(row=row, column=1).value = row - 4
            ws.cell(row=row, column=2).value = nama
            break
        elif nama_cell == nama:
            break
        row += 1

    # Cari kolom tanggal hari ini
    for col in range(3, 34):
        header = ws.cell(row=4, column=col).value
        if header and tanggal_str in header:
            if ws.cell(row=row, column=col).value:
                return False  # Sudah dicatat

            jam_kehadiran = datetime.now().strftime("%H:%M")
            cell = ws.cell(row=row, column=col)
            cell.value = jam_kehadiran
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

            jam_obj = datetime.strptime(jam_kehadiran, "%H:%M").time()
            if jam_obj <= JAM_TEPAT_WAKTU:
                nilai = ws.cell(row=row, column=34).value or 0
                ws.cell(row=row, column=34).value = nilai + 1
                cell.fill = fill_hijau
            else:
                nilai = ws.cell(row=row, column=35).value or 0
                ws.cell(row=row, column=35).value = nilai + 1
                cell.fill = fill_merah

            ws.cell(row=row, column=34).border = thin_border
            ws.cell(row=row, column=35).border = thin_border

            # Tambahkan border di seluruh baris
            for c in range(1, 36):
                ws.cell(row=row, column=c).border = thin_border

            dicatat_hari_ini.add(key_hari_ini)
            wb.save(EXCEL_PATH)

            # === Kirim Telegram Sekali Sehari ===
            tanggal_hari_ini = now.strftime("%Y-%m-%d")
            key_telegram = f"{nama}-{tanggal_hari_ini}"
            if chat_id and key_telegram not in telegram_dicatat_hari_ini:
                status = "Tepat Waktu" if jam_obj <= JAM_TEPAT_WAKTU else "Terlambat"
                pesan = f"Halo, {nama} telah melakukan absensi pada pukul {jam_kehadiran}. Status: {status}."
                kirim_telegram(chat_id, pesan)
                telegram_dicatat_hari_ini.add(key_telegram)

            return True  # Dicatat dan notifikasi dikirim jika ada

    return False  # Tidak berhasil dicatat


# === MediaPipe Face Detection & FaceMesh ===
mp_face_detection = mp.solutions.face_detection
mp_drawing = mp.solutions.drawing_utils
face_detector = mp_face_detection.FaceDetection(model_selection=1, min_detection_confidence=0.6)

mp_face_mesh = mp.solutions.face_mesh
face_mesh = mp_face_mesh.FaceMesh(static_image_mode=False, max_num_faces=1, refine_landmarks=True)

blink_buffer = deque(maxlen=5)
blink_timestamps = {}

def calculate_ear(eye_landmarks):
    A = np.linalg.norm(np.array(eye_landmarks[1]) - np.array(eye_landmarks[5]))
    B = np.linalg.norm(np.array(eye_landmarks[2]) - np.array(eye_landmarks[4]))
    C = np.linalg.norm(np.array(eye_landmarks[0]) - np.array(eye_landmarks[3]))
    ear = (A + B) / (2.0 * C)
    return ear

def detect_blink(frame, name):
    h, w, _ = frame.shape
    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    results = face_mesh.process(rgb)
    if not results.multi_face_landmarks:
        return False
    for face_landmarks in results.multi_face_landmarks:
        mesh_points = [(int(p.x * w), int(p.y * h)) for p in face_landmarks.landmark]
        left_eye_idx = [362, 385, 387, 263, 373, 380]
        right_eye_idx = [33, 160, 158, 133, 153, 144]
        left_eye = [mesh_points[i] for i in left_eye_idx]
        right_eye = [mesh_points[i] for i in right_eye_idx]
        left_ear = calculate_ear(left_eye)
        right_ear = calculate_ear(right_eye)
        ear = (left_ear + right_ear) / 2.0
        if ear < 0.2:
            blink_timestamps[name] = time.time()
            return True
    return False

def recent_blink(name, threshold=2.0):
    return name in blink_timestamps and (time.time() - blink_timestamps[name]) < threshold

def simpan_gambar_wajah(face_img, name, tanggal):
    folder_path = os.path.join("absensi_captured", name)
    os.makedirs(folder_path, exist_ok=True)
    filename = f"{tanggal}.jpg"  # Hanya satu foto per tanggal
    path = os.path.join(folder_path, filename)
    cv2.imwrite(path, face_img)

def kirim_telegram(chat_id, message):
    token = '7982276917:AAHMfjQ6XYubVUWBjTq3-TUAFO-dMf2eF7U'
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = {
        'chat_id': chat_id,
        'text': message
    }
    try:
        requests.post(url, data=payload)
    except Exception as e:
        print("Gagal kirim Telegram:", e)

def get_chat_id(nama_siswa):
    try:
        with open('chat_ids.json', 'r') as f:
            daftar_chat = json.load(f)
        return daftar_chat.get(nama_siswa, None)
    except FileNotFoundError:
        print("File chat_ids.json tidak ditemukan.")
        return None

def play_success_sound():
    try:
        pygame.mixer.init()
        pygame.mixer.music.load('tools/pop.mp3')  # Ganti sesuai nama file
        pygame.mixer.music.play()
    except Exception as e:
        print("Gagal memutar suara:", e)


# === Camera Thread ===
def camera_reader():
    global running
    cap = cv2.VideoCapture(0)

    def get_max_resolution(cap):
        resolutions = [
            (3840, 2160), (2560, 1440), (1920, 1080), (1280, 720),
            (1024, 576), (800, 600), (640, 480)
        ]
        for width, height in resolutions:
            cap.set(cv2.CAP_PROP_FRAME_WIDTH, width)
            cap.set(cv2.CAP_PROP_FRAME_HEIGHT, height)
            actual_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            actual_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            if actual_width == width and actual_height == height:
                return (width, height)
        return (640, 480)  # fallback

    width, height = get_max_resolution(cap)

    while running:
        ret, frame = cap.read()
        if not ret:
            continue
        frame = cv2.flip(frame, 1)
        with frame_lock:
            latest_frame[0] = frame.copy()
        time.sleep(0.005)

    cap.release()


# === Face Recognition Thread ===
def face_recognition_worker():
    global last_detection, notifikasi_text, notifikasi_timer
    frame_count = 0
    while True:
        with frame_lock:
            frame = latest_frame[0]
        if frame is None:
            time.sleep(0.01)
            continue

        if frame_count % PROCESS_EVERY_N_FRAMES == 0:
            try:
                rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                results = face_detector.process(rgb)
                result = []

                if results.detections:
                    for detection in results.detections:
                        bboxC = detection.location_data.relative_bounding_box
                        ih, iw, _ = frame.shape
                        x1 = int(bboxC.xmin * iw)
                        y1 = int(bboxC.ymin * ih)
                        w_box = int(bboxC.width * iw)
                        h_box = int(bboxC.height * ih)

                        # Perbesar bounding box
                        scale = 1.5
                        cx = x1 + w_box // 2
                        cy = y1 + h_box // 2
                        w_scaled = int(w_box * scale)
                        h_scaled = int(h_box * scale)

                        x1 = max(0, cx - w_scaled // 2)
                        y1 = max(0, cy - h_scaled // 2)
                        x2 = min(iw, cx + w_scaled // 2)
                        y2 = min(ih, cy + h_scaled // 2)

                        face = frame[y1:y2, x1:x2]
                        if face.size == 0 or face.shape[0] < 60 or face.shape[1] < 60:
                            continue

                        face_resized = cv2.resize(face, (160, 160), interpolation=cv2.INTER_CUBIC)
                        tensor = torch.from_numpy(face_resized / 255.).permute(2, 0, 1).float()
                        tensor = normalize(tensor).unsqueeze(0).to(device)

                        with torch.no_grad():
                            embedding = resnet(tensor).cpu().numpy()

                        pred = model.predict(embedding)[0]
                        prob = model.predict_proba(embedding)[0]
                        confidence = np.max(prob)
                        name = label_encoder.inverse_transform([pred])[0] if confidence >= UNKNOWN_THRESHOLD else "Gk Kenal"

                        result.append((x1, y1, x2, y2, name, confidence))

                        # Deteksi pengenalan wajah
                        if name != "Gk Kenal":
                            now = time.time()
                            if name not in recognized_since:
                                recognized_since[name] = now

                            duration = now - recognized_since[name]
                            if duration >= 1.0:
                                if detect_blink(frame, name) and recent_blink(name):
                                    jam_datang = datetime.now().strftime("%H:%M")
                                    chat_id = get_chat_id(name)
                                    berhasil = tulis_kehadiran_excel(name, jam_datang, chat_id=chat_id)
                                    if berhasil:
                                        notifikasi_text = f"{name} berhasil dicatat!"
                                        notifikasi_timer = time.time()
                                        threading.Thread(target=play_success_sound, daemon=True).start()
                                        # Simpan foto wajah sekali per hari
                                        tanggal_hari_ini = datetime.now().strftime("%Y-%m-%d")
                                        key_foto = f"{name}-{tanggal_hari_ini}"
                                        if key_foto not in foto_dicatat_hari_ini:
                                            simpan_gambar_wajah(face_resized, name, tanggal_hari_ini)
                                            foto_dicatat_hari_ini.add(key_foto)

                                    # Hindari reset recognized_since agar tidak mendeteksi ulang terus-menerus
                        else:
                            # Reset jika wajah tak dikenal
                            if name in recognized_since:
                                recognized_since.pop(name, None)

                last_detection = result
            except Exception as e:
                print("Recognition error:", e)

        frame_count += 1
        time.sleep(0.01)

# === Display Thread ===
def display_loop():
    global running, notifikasi_text, notifikasi_timer
    window_name = 'Face Recognition (Absensi)'
    cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
    cv2.setWindowProperty(window_name, cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)

    prev_time = time.time()
    fps = 0
    fps_display = 0
    fps_timer = time.time()

    while running:
        with frame_lock:
            frame = latest_frame[0].copy() if latest_frame[0] is not None else None
        if frame is None:
            continue

        # Hitung FPS
        current_time = time.time()
        elapsed = current_time - prev_time
        prev_time = current_time
        fps = 1 / elapsed if elapsed > 0 else 0
        if (current_time - fps_timer) >= 0.5:
            fps_display = fps
            fps_timer = current_time

        # Tampilkan deteksi wajah
        for x1, y1, x2, y2, name, conf in last_detection:
            box_color = (0, 255, 0) if name != "Gk Kenal" else (0, 0, 255)
            text_color = (0, 255, 255)
            outline_color = (0, 0, 0)

            cv2.rectangle(frame, (x1, y1), (x2, y2), box_color, 2)
            label = f"{name} ({conf:.2f})"
            font_scale = 1.2
            thickness_outline = 10
            thickness_text = 2

            cv2.putText(frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, font_scale, outline_color, thickness_outline)
            cv2.putText(frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, font_scale, text_color, thickness_text)

        # Tambahkan teks FPS
        fps_text = f"FPS: {fps_display:.2f}"
        cv2.putText(frame, fps_text, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 0), 3)
        cv2.putText(frame, fps_text, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2)

        # ======== Notifikasi Kuning di Kiri Bawah ========
        if notifikasi_text and (time.time() - notifikasi_timer) < 3:
            h, w, _ = frame.shape
            text_pos = (20, h - 20)  # pojok kiri bawah

            font_scale = 0.9
            thickness_outline = 4
            thickness_text = 2

            # Outline hitam
            cv2.putText(frame, notifikasi_text, text_pos, cv2.FONT_HERSHEY_SIMPLEX, font_scale, (0, 0, 0), thickness_outline, cv2.LINE_AA)
            # Teks utama kuning
            cv2.putText(frame, notifikasi_text, text_pos, cv2.FONT_HERSHEY_SIMPLEX, font_scale, (0, 255, 255), thickness_text, cv2.LINE_AA)

        # Tampilkan frame
        cv2.imshow(window_name, frame)

        if cv2.getWindowProperty(window_name, cv2.WND_PROP_VISIBLE) < 1:
            running = False
            break

        key = cv2.waitKey(1) & 0xFF
        if key == 27:  # ESC
            running = False
            break

    cv2.destroyAllWindows()




# === Run Threads ===
threading.Thread(target=camera_reader, daemon=True).start()
threading.Thread(target=face_recognition_worker, daemon=True).start()
display_loop()
