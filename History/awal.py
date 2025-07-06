import cv2
import torch
import joblib
import numpy as np
import threading
import time
import os
from datetime import datetime
from facenet_pytorch import MTCNN, InceptionResnetV1
from torchvision import transforms
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# === Load Model & Encoder ===
model_data = joblib.load('model.pkl')
model = model_data['model']
label_encoder = model_data['label_encoder']

# === MTCNN & ResNet ===
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
mtcnn = MTCNN(keep_all=True, device=device)
resnet = InceptionResnetV1(pretrained='vggface2').eval().to(device)
normalize = transforms.Normalize([0.5], [0.5])

# === Shared State ===
latest_frame = [None]
frame_lock = threading.Lock()
last_detection = []
already_marked = set()
dicatat_hari_ini = set()
foto_dicatat_hari_ini = set()
running = True
recent_notifications = {}
NOTIFICATION_DURATION = 3  # detik

# === Config ===
PROCESS_EVERY_N_FRAMES = 2
UNKNOWN_THRESHOLD = 0.60
EXCEL_PATH = 'attendance.xlsx'
JAM_TEPAT_WAKTU = datetime.strptime("07:30", "%H:%M").time()

# === Excel Absensi ===
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
fill_hijau = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fill_merah = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def tulis_kehadiran_excel(nama, jam_datang):
    now = datetime.now()
    bulan_tahun = now.strftime("%B %Y")  # Contoh: "June 2025"
    hari_tanggal = now.strftime("%a/%d")
    tanggal_str = now.strftime("%d")
    key_hari_ini = f"{nama}-{hari_tanggal}"

    if key_hari_ini in dicatat_hari_ini:
        return

    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = bulan_tahun
        wb.save(EXCEL_PATH)

    wb = load_workbook(EXCEL_PATH)
    if bulan_tahun not in wb.sheetnames:
        ws = wb.create_sheet(title=bulan_tahun)
    else:
        ws = wb[bulan_tahun]
    # Cek apakah header sudah dibuat
    if not ws.cell(row=1, column=3).value:  # Jika belum ada tulisan "Face Recognition"
        # Header baris 1
        ws.merge_cells('C1:AG1')
        header_cell = ws.cell(row=1, column=3)
        header_cell.value = "Face Recognition"
        header_cell.font = Font(size=26, bold=True)
        header_cell.alignment = Alignment(horizontal='center')
    
        # Bulan
        ws['A2'] = bulan_tahun
        ws['A2'].font = Font(size=14, bold=True)
    
        # Header kolom NO dan Nama
        ws.merge_cells('A3:A4')
        cell = ws['A3']
        cell.value = "NO"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
    
        ws.merge_cells('B3:B4')
        cell = ws['B3']
        cell.value = "Nama"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
    
        # Header tanggal
        ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=33)
        cell = ws.cell(row=3, column=3)
        cell.value = "Waktu Kedatangan"
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
    
        # Header keterangan
        ws.merge_cells(start_row=3, start_column=34, end_row=3, end_column=35)
        cell = ws.cell(row=3, column=34)
        cell.value = "Keterangan"
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
    
        # Baris tanggal (Sen/01, Sel/02, dst)
        for i in range(31):
            try:
                tanggal = datetime(now.year, now.month, i + 1)
                kolom = i + 3
                cell = ws.cell(row=4, column=kolom)
                cell.value = tanggal.strftime("%a/%d")
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True)
            except:
                break
            
        # Keterangan On Time / Late
        ws.cell(row=4, column=34).value = "On Time"
        ws.cell(row=4, column=35).value = "Late"
        for col in [34, 35]:
            cell = ws.cell(row=4, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True)
    
        # Tambahkan border ke seluruh header
        for row in range(3, 5):
            for col in range(1, 36):
                ws.cell(row=row, column=col).border = thin_border

    # Cek baris untuk nama
    row = 5
    while True:
        nama_cell = ws.cell(row=row, column=2).value
        if nama_cell is None:
            ws.cell(row=row, column=1).value = row - 4
            ws.cell(row=row, column=2).value = nama
            break
        elif nama_cell == nama:
            break
        row += 1

    # Cari kolom untuk tanggal hari ini
    for col in range(3, 34):
        header = ws.cell(row=4, column=col).value
        if header and tanggal_str in header:
            if ws.cell(row=row, column=col).value:
                return  # Sudah dicatat

            jam_kehadiran = datetime.now().strftime("%H:%M")
            cell = ws.cell(row=row, column=col)
            cell.value = jam_kehadiran
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            jam_obj = datetime.strptime(jam_kehadiran, "%H:%M").time()

            if jam_obj <= JAM_TEPAT_WAKTU:
                nilai = ws.cell(row=row, column=34).value or 0
                ws.cell(row=row, column=34).value = nilai + 1
                cell.fill = fill_hijau
            else:
                nilai = ws.cell(row=row, column=35).value or 0
                ws.cell(row=row, column=35).value = nilai + 1
                cell.fill = fill_merah

            ws.cell(row=row, column=34).border = thin_border
            ws.cell(row=row, column=35).border = thin_border
            break

    # Tambahkan border ke seluruh baris baru
    for col in range(1, 36):
        ws.cell(row=row, column=col).border = thin_border

    wb.save(EXCEL_PATH)
    dicatat_hari_ini.add(key_hari_ini)
    recent_notifications[nama] = time.time()




def simpan_gambar_wajah(face_img, name, tanggal):
    folder_path = os.path.join("absensi_captured", name)
    os.makedirs(folder_path, exist_ok=True)
    filename = f"{tanggal}.jpg"  # Hanya satu foto per tanggal
    path = os.path.join(folder_path, filename)
    cv2.imwrite(path, face_img)

# === Camera Thread ===
def camera_reader():
    global running
    cap = cv2.VideoCapture(0)

    def get_max_resolution(cap):
        resolutions = [
            (3840, 2160), (2560, 1440), (1920, 1080), (1280, 720),
            (1024, 576), (800, 600), (640, 480)
        ]
        for width, height in resolutions:
            cap.set(cv2.CAP_PROP_FRAME_WIDTH, width)
            cap.set(cv2.CAP_PROP_FRAME_HEIGHT, height)
            actual_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            actual_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            if actual_width == width and actual_height == height:
                return (width, height)
        return (640, 480)  # fallback

    width, height = get_max_resolution(cap)

    while running:
        ret, frame = cap.read()
        if not ret:
            continue
        frame = cv2.flip(frame, 1)
        with frame_lock:
            latest_frame[0] = frame.copy()
        time.sleep(0.005)

    cap.release()


# === Face Recognition Thread ===
def face_recognition_worker():
    global last_detection
    frame_count = 0
    while True:
        with frame_lock:
            frame = latest_frame[0]
        if frame is None:
            time.sleep(0.01)
            continue

        if frame_count % PROCESS_EVERY_N_FRAMES == 0:
            try:
                rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                boxes, _ = mtcnn.detect(rgb)
                result = []

                if boxes is not None:
                    for box in boxes:
                        x1, y1, x2, y2 = [int(b) for b in box]

                        face = frame[y1:y2, x1:x2]
                        if face.size == 0 or face.shape[0] < 60 or face.shape[1] < 60:
                            continue
                        
                        face = cv2.resize(face, (160, 160), interpolation=cv2.INTER_CUBIC)
                        tensor = torch.from_numpy(face / 255.).permute(2, 0, 1).float()
                        tensor = normalize(tensor).unsqueeze(0).to(device)

                        with torch.no_grad():
                            embedding = resnet(tensor).cpu().numpy()

                        pred = model.predict(embedding)[0]
                        prob = model.predict_proba(embedding)[0]
                        confidence = np.max(prob)
                        name = label_encoder.inverse_transform([pred])[0] if confidence >= UNKNOWN_THRESHOLD else "Gk Kenal"

                        result.append((x1, y1, x2, y2, name, confidence))

                        if name != "Gk Kenal":
                            jam_datang = datetime.now().strftime("%H:%M")
                            tulis_kehadiran_excel(name, jam_datang)

                            tanggal_hari_ini = datetime.now().strftime("%Y-%m-%d")
                            key_foto = f"{name}-{tanggal_hari_ini}"

                            if key_foto not in foto_dicatat_hari_ini:
                                face_bgr = cv2.resize(face, (160, 160))
                                simpan_gambar_wajah(face_bgr, name, tanggal_hari_ini)
                                foto_dicatat_hari_ini.add(key_foto)

                last_detection = result

            except Exception as e:
                print("Recognition error:", e)

        frame_count += 1
        time.sleep(0.01)

# === Display Thread ===
def display_loop():
    global running
    window_name = 'Face Recognition (Absensi)'
    cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
    cv2.setWindowProperty(window_name, cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)  # Langsung fullscreen

    prev_time = time.time()
    fps = 0
    fps_display = 0
    fps_timer = time.time()

    while running:
        with frame_lock:
            frame = latest_frame[0].copy() if latest_frame[0] is not None else None
        if frame is None:
            continue

        # Hitung FPS
        current_time = time.time()
        elapsed = current_time - prev_time
        prev_time = current_time
        fps = 1 / elapsed if elapsed > 0 else 0
        if (current_time - fps_timer) >= 0.5:
            fps_display = fps
            fps_timer = current_time

        # Tampilkan deteksi wajah
        for x1, y1, x2, y2, name, conf in last_detection:
            box_color = (0, 255, 0) if name != "Gk Kenal" else (0, 0, 255)
            text_color = (0, 255, 255)
            outline_color = (0, 0, 0)

            cv2.rectangle(frame, (x1, y1), (x2, y2), box_color, 2)
            label = f"{name} ({conf:.2f})"
            font_scale = 1.2
            thickness_outline = 10
            thickness_text = 2

            cv2.putText(frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, font_scale, outline_color, thickness_outline)
            cv2.putText(frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, font_scale, text_color, thickness_text)

        # Tambahkan teks FPS
        fps_text = f"FPS: {fps_display:.2f}"
        cv2.putText(frame, fps_text, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 0), 3)         # Outline hitam
        cv2.putText(frame, fps_text, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2)     # Teks utama

        # === Notifikasi "tercatat" ===
        notif_y = 60
        for name, notif_time in list(recent_notifications.items()):
            if time.time() - notif_time <= NOTIFICATION_DURATION:
                notif_text = f"{name} tercatat!"
                cv2.putText(frame, notif_text, (10, notif_y), cv2.FONT_HERSHEY_SIMPLEX,
                            1.0, (0, 0, 0), 4)  # Outline hitam
                cv2.putText(frame, notif_text, (10, notif_y), cv2.FONT_HERSHEY_SIMPLEX,
                            1.0, (0, 255, 0), 2)  # Teks hijau terang
                notif_y += 40
            else:
                del recent_notifications[name]

        # Tampilkan frame
        cv2.imshow(window_name, frame)

        # Cek jika jendela ditutup manual
        if cv2.getWindowProperty(window_name, cv2.WND_PROP_VISIBLE) < 1:
            running = False
            break

        key = cv2.waitKey(1) & 0xFF
        if key == 27:  # ESC
            running = False
            break

    cv2.destroyAllWindows()




# === Run Threads ===
threading.Thread(target=camera_reader, daemon=True).start()
threading.Thread(target=face_recognition_worker, daemon=True).start()
display_loop()
