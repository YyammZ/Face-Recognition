import streamlit as st
import pandas as pd
from openpyxl import load_workbook

st.set_page_config(page_title="Dashboard Absensi Wajah", layout="wide")
st.title("📸 Dashboard Absensi Wajah")

EXCEL_PATH = "attendance.xlsx"

# === Fungsi Ambil Daftar Sheet (Bulan) ===
def get_sheet_names(path):
    try:
        wb = load_workbook(path, read_only=True)
        return wb.sheetnames
    except:
        return []

# === Fungsi Baca Data Absensi per Sheet ===
def load_absensi_bulan(sheet_name):
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name, header=3)
        return df
    except:
        return None

# === Dropdown Pemilihan Bulan ===
sheet_names = get_sheet_names(EXCEL_PATH)

if not sheet_names:
    st.error("File attendance.xlsx tidak ditemukan atau belum memiliki sheet.")
else:
    selected_sheet = st.selectbox("📅 Pilih Bulan Absensi", sheet_names[::-1])  # Terbaru di atas
    df = load_absensi_bulan(selected_sheet)

    if df is not None:
        st.subheader(f"📋 Data Absensi - {selected_sheet}")
        st.dataframe(df.style.set_properties(**{
            'text-align': 'center',
            'border': '1px solid #999'
        }))
    else:
        st.warning(f"Gagal memuat data dari bulan: {selected_sheet}")
